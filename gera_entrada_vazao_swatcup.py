import os
import pandas as pd
from datetime import datetime
from tqdm import tqdm   # Acrescentado tqdm para barra de progresso
import openpyxl
from openpyxl.styles import Font, Alignment

def ler_csv(arquivo_csv_ou_txt, cod_estacao, dt_inicio, dt_fim, encoding="utf-8"):
    """Lê o csv."""
    if arquivo_csv_ou_txt.endswith('.csv'):
        df = pd.read_csv(arquivo_csv_ou_txt, encoding=encoding , engine='python')
    elif arquivo_csv_ou_txt.endswith('.txt'):
        df = pd.read_csv(arquivo_csv_ou_txt, sep='\t', encoding=encoding, engine='python')
    else:
        raise ValueError("Arquivo deve ser .csv ou .txt")
    
    # Corrigir nomes das colunas para remover aspas e facilitar o acesso
    df.columns = [col.strip().replace('"', '') for col in df.columns]
       
    # print(df.columns)   
            
    if "Cod.estacao" in df.columns:
        df.rename(columns={"Cod.estacao": "cod_estacao"}, inplace=True)
        
    if "Cod_estacao" in df.columns:
        df.rename(columns={"Cod_estacao": "cod_estacao"}, inplace=True)
            
    if "cod_estacao" not in df.columns:
        raise ValueError("A coluna 'cod_estacao' não foi encontrada no arquivo.")
    
    df = df[df["cod_estacao"] == cod_estacao]
    
    df = df[(df["Data"] >= dt_inicio) & (df["Data"] <= dt_fim)]
    
    return df

def criar_arquivo_dados_vazao_SWATCUP_geral(df, var_nome: str, nome_arquivo: str, texto: str = "FLOW_OUT"):
    """Cria um arquivo de dados de vazão formatado."""

    df['n'] = df.index + 1
    df[var_nome] = df[var_nome].round(2)

    with open(f"{nome_arquivo}.txt", 'w') as f:

        for index, row in df.iterrows():
            data = row['Data']
            data = data.strftime("%d_%m_%Y")
            f.write(f"\t{texto}_{data}\t{row['n']}\t{row[var_nome]}\n")
        print(f"Arquivo '{nome_arquivo}' criado com sucesso.")


def criar_arquivo_dados_vazao_SWATCUP_condicionamento_dia(
   caminho_salvar_arquivos, df, var_nome: str, nome_arquivo: int, condicao: str, texto: str = "FLOW_OUT", proporcao_cal: float = 0.7, dt_modelo=None
):
    """
    Cria um arquivo de dados de vazão diário formatado e faz o split em anos completos.
    'condicao' pode ser 'cal' (calibração) ou 'val' (validação).
    'proporcao_cal' permite inverter: 0.3 para calibrar com 30% inicial, 0.7 padrão para 70% inicial.
    'dt_modelo' define o ANO, MÊS e DIA onde a contagem de dias começa.
      - None (default): usa ano, mês e dia do primeiro dado disponível
      - int: assume como ano, mês=1, dia=1
      - (ano, mes, dia): define ano, mês e dia de início da contagem de n
    """

    # Nome do arquivo int to str
    nome_arquivo = str(nome_arquivo)

    df['Data'] = pd.to_datetime(df['Data'])
    df[var_nome] = df[var_nome].round(2)
    df = df[df[var_nome].notna()]

    # Encontrar o cutoff em proporção de linhas
    cutoff_idx = int(len(df) * proporcao_cal)
    cutoff_data = df.iloc[cutoff_idx]['Data']
    cutoff_year = cutoff_data.year

    if condicao == "cal":
        # Calibração: até o último dia do ano do cutoff
        fim_cal = pd.Timestamp(year=cutoff_year, month=12, day=31)
        df_sel = df[df['Data'] <= fim_cal]
        print("Calibração:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "val":
        # Validação: a partir do primeiro dia do ano seguinte ao cutoff
        ini_val = pd.Timestamp(year=cutoff_year + 1, month=1, day=1)
        df_sel = df[df['Data'] >= ini_val]
        print("Validação:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    else:
        raise ValueError("condicao deve ser 'cal' ou 'val'")

    # --- Lógica para dt_modelo
    if dt_modelo is None:
        ano_inicio = df_sel['Data'].iloc[0].year
        mes_inicio = df_sel['Data'].iloc[0].month
        dia_inicio = df_sel['Data'].iloc[0].day
    elif isinstance(dt_modelo, int):
        ano_inicio = dt_modelo
        mes_inicio = 1
        dia_inicio = 1
    elif isinstance(dt_modelo, (tuple, list)) and len(dt_modelo) == 3:
        ano_inicio, mes_inicio, dia_inicio = dt_modelo
    else:
        raise ValueError("dt_modelo deve ser None, int ou (ano, mes, dia)")

    dt_inicio = pd.Timestamp(year=ano_inicio, month=mes_inicio, day=dia_inicio)

    # Calcular n para cada linha baseada em dt_modelo
    df_sel = df_sel.copy()
    df_sel['n'] = (df_sel['Data'] - dt_inicio).dt.days + 1

    # --- Gera o arquivo de saída
    caminho_salvar_arquivos = os.path.join(caminho_salvar_arquivos, "VAZOES_TXT")
    
    if not os.path.exists(caminho_salvar_arquivos):
        os.makedirs(caminho_salvar_arquivos)

    with open(f"{caminho_salvar_arquivos}/{nome_arquivo}_{condicao}_{proporcao_cal}_dia.txt", 'w') as f:
        for _, row in df_sel.iterrows():
            data = row['Data'].strftime("%d_%m_%Y")
            f.write(f"{row['n']}\t{texto}_{data}\t{row[var_nome]}\n")
        print(f"Arquivo '{nome_arquivo}_{condicao}_{proporcao_cal}_dia.txt' criado com sucesso.")
        
def dataframe_vazao_formatado_condicionamento_dia(
    df, var_nome: str, nome_arquivo: int, condicao: str, texto: str = "FLOW_OUT", proporcao_cal: float = 0.7, dt_modelo=None
):
    """
    Cria um dataframe de dados de vazão diário formatado e faz o split em anos completos.
    'condicao' pode ser 'cal' (calibração), 'val' (validação) ou 'todas' (todas as datas).
    'proporcao_cal' permite inverter: 0.3 para calibrar com 30% inicial, 0.7 padrão para 70% inicial.
    'dt_modelo' define o ANO, MÊS e DIA onde a contagem de dias começa.
      - None (default): usa ano, mês e dia do primeiro dado disponível
      - int: assume como ano, mês=1, dia=1
      - (ano, mes, dia): define ano, mês e dia de início da contagem de n
    """

    df = df.copy()
    nome_arquivo = str(nome_arquivo)

    df['Data'] = pd.to_datetime(df['Data'])
    # Agrupa para garantir datas únicas, caso haja duplicidade
    daily = df.groupby(df['Data'])[[var_nome]].mean().reset_index()
    daily[var_nome] = daily[var_nome].round(2)
    daily = daily[daily[var_nome].notna()]

    # Encontrar o cutoff em proporção de linhas
    if condicao == "todas":
        pass
    else:
        cutoff_idx = int(len(daily) * proporcao_cal)
        cutoff_data = daily.iloc[cutoff_idx]['Data']
        cutoff_year = cutoff_data.year

    if condicao == "cal":
        fim_cal = pd.Timestamp(year=cutoff_year, month=12, day=31)
        df_sel = daily[daily['Data'] <= fim_cal]
        print("Calibração:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "val":
        ini_val = pd.Timestamp(year=cutoff_year + 1, month=1, day=1)
        df_sel = daily[daily['Data'] >= ini_val]
        print("Validação:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "todas":
        df_sel = daily
        print("Todas as datas:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    else:
        raise ValueError("condicao deve ser 'cal' ou 'val'")

    # --- Lógica para dt_modelo
    if dt_modelo is None:
        ano_inicio = df_sel['Data'].iloc[0].year
        mes_inicio = df_sel['Data'].iloc[0].month
        dia_inicio = df_sel['Data'].iloc[0].day
    elif isinstance(dt_modelo, int):
        ano_inicio = dt_modelo
        mes_inicio = 1
        dia_inicio = 1
    elif isinstance(dt_modelo, (tuple, list)) and len(dt_modelo) == 3:
        ano_inicio, mes_inicio, dia_inicio = dt_modelo
    else:
        raise ValueError("dt_modelo deve ser None, int ou (ano, mes, dia)")

    # Calcular n para cada linha baseada em dt_modelo
    def calcula_n(row):
        return (row['Data'] - pd.Timestamp(ano_inicio, mes_inicio, dia_inicio)).days + 1

    df_sel = df_sel.copy()
    df_sel['n'] = df_sel.apply(calcula_n, axis=1)
    
    return df_sel

def salvar_multiplos_dataframes_vazao_excel(
    lista_dataframes, lista_labels, nome_arquivo: str, cod_estacao: int, num_flow_out: int, texto: str = "FLOW_OUT", var_nome: str = "Vazao"
):
    """
    Salva múltiplos dataframes formatados em um arquivo Excel lado a lado.
    
    Args:
        lista_dataframes: Lista de DataFrames retornados pela função dataframe_vazao_formatado_condicionamento_dia
        lista_labels: Lista de labels para cada dataframe (ex: ['cal_0.7', 'val_0.7', 'cal_0.3', 'val_0.3'])
        nome_arquivo: Nome base do arquivo Excel (sem extensão)
        cod_estacao: Código da estação
        num_flow_out: Número do FLOW_OUT para o cabeçalho
        texto: Prefixo dos dados (padrão "FLOW_OUT")
        var_nome: Nome da coluna de vazão (padrão "Vazao")
    """
    
    # Criar workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Estacao_{cod_estacao}_multiplos"
    
    num_secoes = len(lista_dataframes)
    
    # Cabeçalhos das seções
    headers = ['N','FLOW_OUT']
    # headers = ['Estação:', 'FLOW_OUT:', 'N:']
    # valores_header = [cod_estacao, num_flow_out, 0]  # N será preenchido dinamicamente
    
    # Escrever cabeçalhos dinamicamente
    for col in range(num_secoes):
        
        N = len(lista_dataframes[col])
        
        # Cabeçalho da seção com label específico
        cell_header = ws.cell(row=1, column=col*4 + 2, value=f"Seção {lista_labels[col]}")
        cell_header.font = Font(bold=True)
        
        # Cabeçalho da seção
        # cell_header = ws.cell(row=2, column=col*4 + 1, value=headers[0])
        # cell_header.font = Font(bold=True)
        
        # Valor do cabeçalho
        # cell_value = ws.cell(row=2, column=col*4 + 2, value=valores_header[0])
        # cell_value.font = Font(bold=True)
        
        # Cabeçalho das colunas de dados
        
        ws.cell(row=2, column=col*4 + 1, value=headers[0] + f": {N}")
        ws.cell(row=2, column=col*4 + 2, value=headers[1] + f": {num_flow_out}")
        ws.cell(row=2, column=col*4 + 3, value='Vazão')
        
        # ws.cell(row=4, column=col*4 + 1, value=headers[0].replace(':', ''))
        # ws.cell(row=4, column=col*4 + 2, value=headers[1].replace(':', ''))
        # ws.cell(row=4, column=col*4 + 3, value=headers[2].replace(':', ''))
        # ws.cell(row=4, column=col*4 + 4, value='Vazão')
    
    # Escrever dados de cada dataframe em sua seção correspondente
    max_linhas = max(len(df) for df in lista_dataframes) if lista_dataframes else 0
    
    for row_idx in range(max_linhas):
        row_num = row_idx + 3  # Começar na linha 3
        
        for secao, df_formatado in enumerate(lista_dataframes):
            col_base = secao * 4  # Cada seção ocupa 4 colunas
            
            if row_idx < len(df_formatado):
                row = df_formatado.iloc[row_idx]
                data_formatada = row['Data'].strftime("%d_%m_%Y")
                flow_out_nome = f"{texto}_{data_formatada}"
                
                ws.cell(row=row_num, column=col_base + 1, value=row['n'])
                ws.cell(row=row_num, column=col_base + 2, value=flow_out_nome)
                ws.cell(row=row_num, column=col_base + 3, value=row[var_nome])
                
                # ws.cell(row=row_num, column=col_base + 1, value=cod_estacao)
                # ws.cell(row=row_num, column=col_base + 2, value=flow_out_nome)
                # ws.cell(row=row_num, column=col_base + 3, value=row['n'])
                # ws.cell(row=row_num, column=col_base + 4, value=row[var_nome])
    
    # Ajustar largura das colunas dinamicamente
    total_colunas = num_secoes * 4
    for col in range(1, total_colunas + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Salvar arquivo
    nome_arquivo_excel = f"{nome_arquivo}_multiplos.xlsx"
    wb.save(nome_arquivo_excel)
    print(f"Arquivo Excel '{nome_arquivo_excel}' criado com sucesso.")

def salvar_todas_estacoes_excel_multiplas_abas(
    caminho_salvar: str, dados_todas_estacoes, nome_arquivo: str = "todas_estacoes", texto: str = "FLOW_OUT", var_nome: str = "Vazao"
):
    """
    Salva dados de múltiplas estações em um único arquivo Excel com abas separadas.
    
    Args:
        dados_todas_estacoes: Dicionário com dados de cada estação no formato:
            {
                cod_estacao: {
                    'num_flow_out': int,
                    'lista_dataframes': [df_cal_70, df_val_30, df_cal_30, df_val_70],
                    'lista_labels': ['cal_0.7', 'val_0.3', 'cal_0.3', 'val_0.7']
                }
            }
        nome_arquivo: Nome base do arquivo Excel (sem extensão)
        texto: Prefixo dos dados (padrão "FLOW_OUT")
        var_nome: Nome da coluna de vazão (padrão "Vazao")
    """
    
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Remover a planilha padrão
    wb.remove(wb.active)
    
    # Criar uma aba para cada estação
    for cod_estacao, dados_estacao in dados_todas_estacoes.items():
        lista_dataframes = dados_estacao['lista_dataframes']
        lista_labels = dados_estacao['lista_labels']
        num_flow_out = dados_estacao['num_flow_out']
        
        # Criar nova aba
        ws = wb.create_sheet(title=f"Estacao_{cod_estacao}")
        
        num_secoes = len(lista_dataframes)
        headers = ['N', 'FLOW_OUT']
        
        # Escrever cabeçalhos dinamicamente
        for col in range(num_secoes):
            N = len(lista_dataframes[col])
            
            # Mesclar células primeiro (3 colunas)
            ws.merge_cells(start_row=1, start_column=col*4 + 1, end_row=1, end_column=col*4 + 3)
            
            # Cabeçalho da seção com label específico (após mesclar)
            cell_header = ws.cell(row=1, column=col*4 + 1, value=f"Seção {lista_labels[col]}")
            cell_header.font = Font(bold=True)
            cell_header.alignment = Alignment(horizontal='center')
            
            # Cabeçalho das colunas de dados
            ws.cell(row=2, column=col*4 + 1, value=headers[0] + f"= {N}")
            ws.cell(row=2, column=col*4 + 2, value=headers[1] + f"_{num_flow_out}") 
            ws.cell(row=2, column=col*4 + 3, value='Vazão')
        
        # Escrever dados de cada dataframe em sua seção correspondente
        max_linhas = max(len(df) for df in lista_dataframes) if lista_dataframes else 0
        
        for row_idx in range(max_linhas):
            row_num = row_idx + 3  # Começar na linha 3
            
            for secao, df_formatado in enumerate(lista_dataframes):
                col_base = secao * 4  # Cada seção ocupa 4 colunas
                
                if row_idx < len(df_formatado):
                    row = df_formatado.iloc[row_idx]
                    data_formatada = row['Data'].strftime("%d_%m_%Y")
                    flow_out_nome = f"{texto}_{data_formatada}"
                    
                    ws.cell(row=row_num, column=col_base + 1, value=row['n'])
                    ws.cell(row=row_num, column=col_base + 2, value=flow_out_nome)
                    ws.cell(row=row_num, column=col_base + 3, value=row[var_nome])
        
        # Ajustar largura das colunas dinamicamente
        total_colunas = num_secoes * 4
        for col in range(1, total_colunas + 1):
            # Verificar se é uma coluna de FLOW_OUT (coluna 2 de cada seção)
            if (col - 2) % 4 == 0 and col > 1:  # Colunas 2, 6, 10, 14, etc.
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 21.43
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Salvar arquivo
    nome_arquivo_excel = f"{nome_arquivo}.xlsx"
    wb.save(os.path.join(caminho_salvar, nome_arquivo_excel))
    print(f"Arquivo Excel '{nome_arquivo_excel}' criado com sucesso com {len(dados_todas_estacoes)} abas.")

def salvar_dataframe_vazao_excel(
    df_formatado, nome_arquivo: str, cod_estacao: int, num_flow_out: int, condicao: str, proporcao_cal: float = 0.7, texto: str = "FLOW_OUT", var_nome: str = "Vazao", num_secoes: int = 3
):
    """
    Salva o dataframe formatado em um arquivo Excel com layout de colunas repetidas.
    
    Args:
        df_formatado: DataFrame retornado pela função dataframe_vazao_formatado_condicionamento_dia
        nome_arquivo: Nome base do arquivo Excel (sem extensão)
        cod_estacao: Código da estação
        num_flow_out: Número do FLOW_OUT para o cabeçalho
        condicao: 'cal' ou 'val'
        proporcao_cal: Proporção de calibração utilizada
        texto: Prefixo dos dados (padrão "FLOW_OUT")
        var_nome: Nome da coluna de vazão (padrão "Vazao")
        num_secoes: Número de seções repetidas no Excel (padrão 3)
    """
    
    # Criar workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Estacao_{cod_estacao}_{condicao}"
    
    # Contagem de Linhas do df_formatado 
    N = len(df_formatado)
    
    # Preparar dados
    dados = []
    for _, row in df_formatado.iterrows():
        data_formatada = row['Data'].strftime("%d_%m_%Y")
        flow_out_nome = f"{texto}_{data_formatada}"
        dados.append([cod_estacao, flow_out_nome, row['n'], row[var_nome]])
    
    # Cabeçalhos das seções
    headers = ['N','FLOW_OUT']
    
    # Escrever cabeçalhos dinamicamente
    for col in range(num_secoes):
      
        # Cabeçalho das colunas de dados
        ws.cell(row=1, column=col*3 + 1, value=headers[0] + f": {N}")
        ws.cell(row=1, column=col*3 + 2, value=headers[1] + f": {num_flow_out}")
        ws.cell(row=1, column=col*3 + 3, value='Vazão')

    # Escrever dados dinamicamente em todas as seções
    for i, (n, flow_out, vazao) in enumerate(dados):
        row_num = i + 2  # Começar na linha 2
        
        # Para cada seção
        for secao in range(num_secoes):
            col_base = secao * 3 # Cada seção ocupa 4 colunas
            ws.cell(row=row_num, column=col_base + 1, value=n)
            ws.cell(row=row_num, column=col_base + 2, value=flow_out)
            ws.cell(row=row_num, column=col_base + 3, value=vazao)
    
    # Ajustar largura das colunas dinamicamente
    total_colunas = num_secoes * 4
    for col in range(1, total_colunas + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Salvar arquivo
    nome_arquivo_excel = f"{nome_arquivo}_{condicao}_{proporcao_cal}.xlsx"
    wb.save(nome_arquivo_excel)
    print(f"Arquivo Excel '{nome_arquivo_excel}' criado com sucesso.")

def criar_arquivo_dados_vazao_SWATCUP_condicionamento_mes(
    caminho_salvar_arquivos, df, var_nome: str, nome_arquivo: int, condicao: str, texto: str = "FLOW_OUT", proporcao_cal: float = 0.7, dt_modelo=None
):
    """
    Cria um arquivo de dados de vazão formatado e faz o split em anos completos.
    'condicao' pode ser 'cal' (calibração) ou 'val' (validação).
    'proporcao_cal' permite inverter: 0.3 para calibrar com 30% inicial, 0.7 padrão para 70% inicial.
    'dt_modelo' define o ANO e MÊS onde a contagem de meses começa. Pode ser:
      - None (default): usa o ano e mês do primeiro dado disponível
      - int: assume como ano e mês=1
      - (ano, mes): define ano e mês de início da contagem de n
    """

    # Nome do arquivo int to str
    nome_arquivo = str(nome_arquivo)

    df['Data'] = pd.to_datetime(df['Data'])
    monthly = df.groupby(df['Data'].dt.to_period("M"))[[var_nome]].mean().reset_index()
    monthly['Data'] = monthly['Data'].dt.to_timestamp()  # Converte Period para Timestamp (primeiro dia do mês)
    monthly[var_nome] = monthly[var_nome].round(2)
    monthly = monthly[monthly[var_nome].notna()]

    # Encontrar o cutoff em proporção de linhas
    cutoff_idx = int(len(monthly) * proporcao_cal)
    cutoff_data = monthly.iloc[cutoff_idx]['Data']
    cutoff_year = cutoff_data.year

    if condicao == "cal":
        # Calibração: até o último dia do ano do cutoff
        fim_cal = pd.Timestamp(year=cutoff_year, month=12, day=31)
        df_sel = monthly[monthly['Data'] <= fim_cal]
        print("Calibração:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "val":
        # Validação: a partir do primeiro dia do ano seguinte ao cutoff
        ini_val = pd.Timestamp(year=cutoff_year + 1, month=1, day=1)
        df_sel = monthly[monthly['Data'] >= ini_val]
        print("Validação:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    else:
        raise ValueError("condicao deve ser 'cal' ou 'val'")

    # --- Lógica para dt_modelo
    if dt_modelo is None:
        ano_inicio = df_sel['Data'].iloc[0].year
        mes_inicio = df_sel['Data'].iloc[0].month
    elif isinstance(dt_modelo, int):
        ano_inicio = dt_modelo
        mes_inicio = 1
    elif isinstance(dt_modelo, (tuple, list)) and len(dt_modelo) == 2:
        ano_inicio, mes_inicio = dt_modelo
    else:
        raise ValueError("dt_modelo deve ser None, int ou (ano, mes)")

    # Calcular n para cada linha baseada em dt_modelo
    def calcula_n(row):
        return (row['Data'].year - ano_inicio) * 12 + (row['Data'].month - mes_inicio) + 1

    df_sel = df_sel.copy()
    df_sel['n'] = df_sel.apply(calcula_n, axis=1)

    # --- Gera o arquivo de saída
    caminho_salvar_arquivos = os.path.join(caminho_salvar_arquivos, "VAZOES_TXT")
    
    if not os.path.exists(caminho_salvar_arquivos):
        os.makedirs(caminho_salvar_arquivos)

    with open(f"{caminho_salvar_arquivos}/{nome_arquivo}_{condicao}_{proporcao_cal}_mes.txt", 'w') as f:
        for _, row in df_sel.iterrows():
            data = row['Data']
            data_str = data.strftime("%m_%Y")
            f.write(f"{row['n']}\t{texto}_{data_str}\t{row[var_nome]}\n")
        print(f"Arquivo '{nome_arquivo}_{condicao}_{proporcao_cal}_mes.txt' criado com sucesso.")
        
        
def dataframe_vazao_formatado_condicionamento_mes(df, var_nome: str, nome_arquivo: int, condicao: str, texto: str = "FLOW_OUT", proporcao_cal: float = 0.7, dt_modelo=None
):
    """
    Cria um arquivo de dados de vazão formatado e faz o split em anos completos.
    'condicao' pode ser 'cal' (calibração), 'val' (validação) ou 'todas' (todas as datas).
    'proporcao_cal' permite inverter: 0.3 para calibrar com 30% inicial, 0.7 padrão para 70% inicial.
    'dt_modelo' define o ANO e MÊS onde a contagem de meses começa. Pode ser:
      - None (default): usa o ano e mês do primeiro dado disponível
      - int: assume como ano e mês=1
      - (ano, mes): define ano e mês de início da contagem de n
    """

    # Nome do arquivo int to str
    nome_arquivo = str(nome_arquivo)

    df['Data'] = pd.to_datetime(df['Data'])
    monthly = df.groupby(df['Data'].dt.to_period("M"))[[var_nome]].mean().reset_index()
    monthly['Data'] = monthly['Data'].dt.to_timestamp()  # Converte Period para Timestamp (primeiro dia do mês)
    monthly[var_nome] = monthly[var_nome].round(2)
    monthly = monthly[monthly[var_nome].notna()]

    # Encontrar o cutoff em proporção de linhas
    if condicao == "todas":
        pass    
    else:
        cutoff_idx = int(len(monthly) * proporcao_cal)
        cutoff_data = monthly.iloc[cutoff_idx]['Data']
        cutoff_year = cutoff_data.year

    if condicao == "cal":
        # Calibração: até o último dia do ano do cutoff
        fim_cal = pd.Timestamp(year=cutoff_year, month=12, day=31)
        df_sel = monthly[monthly['Data'] <= fim_cal]
        print("Calibração:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "val":
        # Validação: a partir do primeiro dia do ano seguinte ao cutoff
        ini_val = pd.Timestamp(year=cutoff_year + 1, month=1, day=1)
        df_sel = monthly[monthly['Data'] >= ini_val]
        print("Validação:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    elif condicao == "todas":
        df_sel = monthly
        print("Todas as datas:", df_sel["Data"].min(), "→", df_sel["Data"].max())
    else:
        raise ValueError("condicao deve ser 'cal' ou 'val'")

    # --- Lógica para dt_modelo
    if dt_modelo is None:
        ano_inicio = df_sel['Data'].iloc[0].year
        mes_inicio = df_sel['Data'].iloc[0].month
    elif isinstance(dt_modelo, int):
        ano_inicio = dt_modelo
        mes_inicio = 1
    elif isinstance(dt_modelo, (tuple, list)) and len(dt_modelo) == 2:
        ano_inicio, mes_inicio = dt_modelo
    else:
        raise ValueError("dt_modelo deve ser None, int ou (ano, mes)")

    # Calcular n para cada linha baseada em dt_modelo
    def calcula_n(row):
        return (row['Data'].year - ano_inicio) * 12 + (row['Data'].month - mes_inicio) + 1

    df_sel = df_sel.copy()
    df_sel['n'] = df_sel.apply(calcula_n, axis=1)
    
    return df_sel  

def criar_arquivo_dados_vazao_SWATCUP_validacao_txt(df, var_nome, nome_arquivo ):
    """
    Cria um arquivo de dados de vazão formatado.
    """

    with open(f"{nome_arquivo}.txt", 'w') as f:

        for index, row in df.iterrows():
            f.write(f"{row[var_nome]}\n")
        print(f"Arquivo '{nome_arquivo}' criado com sucesso.")


if __name__ == "__main__":
    saida = r"C:\Users\rogerio.siqueira\Documents\DEMANDAS\Análise e entradas de dados Plu-Flu no QSWAT e SWATCUP"
    flow_out = [8, 13, 20] # Lista de estações para processamento
    estacoes = [60476100, 60471200, 60474100]  # Lista de estações para processamento
    datas_inicio = ["1978-01-01", "1990-01-01", "1995-01-01"]
    datas_fim = ["2014-12-31", "2022-12-31", "2022-12-31"]
    
    # Dicionário para armazenar dados de todas as estações
    dados_todas_estacoes_mes = {}
    dados_todas_estacoes_dia = {}
    
    for estacao, data_inicio, data_fim, num_flow_out in zip(estacoes, datas_inicio, datas_fim, flow_out):
        df = ler_csv("FLU_Series_ANA.txt", 
                     estacao, 
                     data_inicio, 
                     data_fim, 
                     encoding='utf-8' # latin1 ou utf-8
        )
        
        # Checar se há nan ou null na coluna 'Vazao'
        if df['Vazao'].isnull().values.any():
            print("Existem valores nulos na coluna 'Vazao'")

        # Mostrar que linhas são
            df_nan = df[df['Vazao'].isnull()]
        
        criar_arquivo_dados_vazao_SWATCUP_condicionamento_dia(saida, df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1,1))
        criar_arquivo_dados_vazao_SWATCUP_condicionamento_dia(saida, df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1,1))

        # print("------------------------------------------------------------------------------")

        criar_arquivo_dados_vazao_SWATCUP_condicionamento_dia(saida, df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1,1))
        criar_arquivo_dados_vazao_SWATCUP_condicionamento_dia(saida, df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1,1))

        print("------------------------------------------------------------------------------")

        criar_arquivo_dados_vazao_SWATCUP_condicionamento_mes(saida, df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1))
        criar_arquivo_dados_vazao_SWATCUP_condicionamento_mes(saida, df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1))

        print("------------------------------------------------------------------------------")

        criar_arquivo_dados_vazao_SWATCUP_condicionamento_mes(saida, df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1))
        criar_arquivo_dados_vazao_SWATCUP_condicionamento_mes(saida, df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1))

        print("------------------------------------------------------------------------------")
        
# Gerar todos os dataframes para as estação - Dia 
        df_cal_70_d = dataframe_vazao_formatado_condicionamento_dia(df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1,1))
        df_val_70_d = dataframe_vazao_formatado_condicionamento_dia(df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1,1))
        df_cal_30_d = dataframe_vazao_formatado_condicionamento_dia(df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1,1))
        df_val_30_d = dataframe_vazao_formatado_condicionamento_dia(df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1,1))
        df_todas_d = dataframe_vazao_formatado_condicionamento_dia(df, "Vazao", estacao, condicao="todas", texto="FLOW_OUT", proporcao_cal=1, dt_modelo=(1978,1,1))

        # Gerar todos os dataframes para a estação -  Mês 
        df_cal_70_m = dataframe_vazao_formatado_condicionamento_mes(df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1))
        df_val_70_m = dataframe_vazao_formatado_condicionamento_mes(df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.7, dt_modelo=(1978,1))
        df_val_30_m = dataframe_vazao_formatado_condicionamento_mes(df, "Vazao", estacao, condicao="val", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1))
        df_cal_30_m = dataframe_vazao_formatado_condicionamento_mes(df, "Vazao", estacao, condicao="cal", texto="FLOW_OUT", proporcao_cal=0.3, dt_modelo=(1978,1))
        df_todas_m = dataframe_vazao_formatado_condicionamento_mes(df, "Vazao", estacao, condicao="todas", texto="FLOW_OUT", proporcao_cal=1, dt_modelo=(1978,1))

        # Lista dos dataframes e seus labels
        lista_dfs_dia = [df_cal_70_d, df_val_70_d, df_cal_30_d, df_val_30_d, df_todas_d] 
        lista_dfs_mes = [df_cal_70_m, df_val_70_m, df_cal_30_m, df_val_30_m, df_todas_m]
        lista_labels = ['cal_0.7', 'val_0.3', 'cal_0.3', 'val_0.7', 'Completo'] # Labels para cada dataframe de acordo com a proporção e condição

        # Armazenar dados da estação no dicionário
        dados_todas_estacoes_mes[estacao] = {
            'num_flow_out': num_flow_out,
            'lista_dataframes': lista_dfs_mes,
            'lista_labels': lista_labels
        }

        dados_todas_estacoes_dia[estacao] = {
            'num_flow_out': num_flow_out,
            'lista_dataframes': lista_dfs_dia,
            'lista_labels': lista_labels
        }

        print("Processamento da estação", estacao, "concluído!")
        print("------------------------------------------------------------------------------")
    
    # Após processar todas as estações, salvar em um único Excel com múltiplas abas
    print("Salvando todas as estações em um único Excel...")
    salvar_todas_estacoes_excel_multiplas_abas(saida, dados_todas_estacoes_mes, "todas_estacoes_vazao_mensal", "FLOW_OUT", "Vazao")
    salvar_todas_estacoes_excel_multiplas_abas(saida, dados_todas_estacoes_dia, "todas_estacoes_vazao_diario", "FLOW_OUT", "Vazao")
    print("Processamento completo finalizado!")